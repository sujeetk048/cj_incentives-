"""
HTML Query Executor
Reads queries from an HTML file and executes them in Snowflake,
outputting results to Excel sheets.
"""
import pandas as pd
import logging
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from snowflake.connector import connect
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import time

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('html_query_executor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def load_env_variables():
    """Load environment variables from .env file"""
    load_dotenv()
    return {
        'account': os.getenv('SNOWFLAKE_ACCOUNT'),
        'user': os.getenv('SNOWFLAKE_USER'),
        'role': os.getenv('SNOWFLAKE_ROLE'),
        'warehouse': os.getenv('SNOWFLAKE_WAREHOUSE'),
        'database': os.getenv('SNOWFLAKE_DATABASE'),
        'schema': os.getenv('SNOWFLAKE_SCHEMA')
    }


def connect_snowflake(config):
    """Connect to Snowflake using configuration"""
    try:
        logger.info("Connecting to Snowflake...")
        conn = connect(
            account=config['account'],
            user=config['user'],
            authenticator='externalbrowser',
            role=config['role'],
            warehouse=config['warehouse'],
            database=config['database'],
            schema=config['schema']
        )
        logger.info("Successfully connected to Snowflake")
        return conn
    except Exception as e:
        logger.error(f"Failed to connect to Snowflake: {str(e)}")
        raise


def read_html_queries(html_file_path):
    """
    Read queries from HTML file
    
    Args:
        html_file_path: Path to HTML file containing query table
        
    Returns:
        DataFrame with columns: query_name, sql_query, sheet_name
    """
    try:
        logger.info(f"Reading queries from HTML file: {html_file_path}")
        
        # Read HTML tables
        tables = pd.read_html(html_file_path)
        
        if not tables:
            raise ValueError("No tables found in HTML file")
        
        # Get the first table
        df = tables[0]
        
        # Normalize column names
        df.columns = df.columns.str.lower().str.replace(' ', '_')
        
        # Check required columns
        required_columns = ['query_name', 'sql_query', 'sheet_name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        logger.info(f"Found {len(df)} queries in HTML file")
        return df
    
    except Exception as e:
        logger.error(f"Failed to read HTML file: {str(e)}")
        raise


def execute_query(conn, query_name, sql_query):
    """
    Execute a single SQL query and return results
    
    Args:
        conn: Snowflake connection
        query_name: Name of the query
        sql_query: SQL query string
        
    Returns:
        Tuple of (DataFrame, execution_time, status)
    """
    start_time = time.time()
    
    try:
        logger.info(f"Executing query: {query_name}")
        
        cursor = conn.cursor()
        cursor.execute(sql_query)
        
        # Fetch results
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        
        df = pd.DataFrame(rows, columns=columns)
        cursor.close()
        
        execution_time = time.time() - start_time
        status = 'success'
        
        logger.info(f"Query '{query_name}' completed in {execution_time:.2f}s - {len(df)} rows")
        
        return df, execution_time, status
    
    except Exception as e:
        execution_time = time.time() - start_time
        status = 'failed'
        error_msg = str(e)
        
        logger.error(f"Query '{query_name}' failed: {error_msg}")
        
        return None, execution_time, status


def write_to_excel(df, sheet_name, excel_file_path, replace_sheet=True):
    """
    Write DataFrame to Excel sheet
    
    Args:
        df: DataFrame to write
        sheet_name: Name of the sheet
        excel_file_path: Path to Excel file
        replace_sheet: If True, replace existing sheet; if False, append
    """
    try:
        # Check if file exists
        file_exists = Path(excel_file_path).exists()
        
        if file_exists and replace_sheet:
            # Load existing workbook
            book = load_workbook(excel_file_path)
            
            # Remove sheet if it exists
            if sheet_name in book.sheetnames:
                del book[sheet_name]
                logger.info(f"Removed existing sheet: {sheet_name}")
            
            # Write new sheet
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        else:
            # Create new file or append
            mode = 'a' if file_exists else 'w'
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode=mode) as writer:
                if file_exists and not replace_sheet:
                    writer.book = load_workbook(excel_file_path)
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Apply formatting
        apply_formatting(excel_file_path, sheet_name)
        
        logger.info(f"Written {len(df)} rows to sheet '{sheet_name}'")
    
    except Exception as e:
        logger.error(f"Failed to write to Excel: {str(e)}")
        raise


def apply_formatting(excel_file_path, sheet_name):
    """Apply formatting to Excel sheet"""
    try:
        book = load_workbook(excel_file_path)
        sheet = book[sheet_name]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Apply header formatting
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        if sheet.max_row >= 1:
            sheet.freeze_panes = 'A2'
        
        book.save(excel_file_path)
    
    except Exception as e:
        logger.warning(f"Failed to apply formatting: {str(e)}")


def main(html_file_path, output_file='incentive_output.xlsx'):
    """
    Main execution function
    
    Args:
        html_file_path: Path to HTML file with queries
        output_file: Output Excel file name
    """
    logger.info("=" * 60)
    logger.info("HTML Query Executor - Starting")
    logger.info("=" * 60)
    
    # Load configuration
    config = load_env_variables()
    
    # Read queries from HTML
    queries_df = read_html_queries(html_file_path)
    
    # Connect to Snowflake
    conn = connect_snowflake(config)
    
    # Execution log
    execution_log = []
    
    try:
        # Process each query
        for index, row in queries_df.iterrows():
            query_name = row['query_name']
            sql_query = row['sql_query']
            sheet_name = row['sheet_name']
            
            logger.info(f"\nProcessing query {index + 1}/{len(queries_df)}: {query_name}")
            
            # Execute query
            df, execution_time, status = execute_query(conn, query_name, sql_query)
            
            # Write to Excel if successful
            if status == 'success' and df is not None:
                write_to_excel(df, sheet_name, output_file, replace_sheet=True)
            
            # Log execution
            execution_log.append({
                'query_name': query_name,
                'sheet_name': sheet_name,
                'status': status,
                'execution_time': f"{execution_time:.2f}s",
                'row_count': len(df) if df is not None else 0,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Save execution log
        log_df = pd.DataFrame(execution_log)
        log_df.to_excel('execution_log.xlsx', index=False)
        logger.info(f"Execution log saved to execution_log.xlsx")
        
        # Summary
        success_count = sum(1 for log in execution_log if log['status'] == 'success')
        logger.info("\n" + "=" * 60)
        logger.info("EXECUTION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Total queries: {len(queries_df)}")
        logger.info(f"Successful: {success_count}")
        logger.info(f"Failed: {len(queries_df) - success_count}")
        logger.info(f"Output file: {output_file}")
        logger.info("=" * 60)
    
    finally:
        conn.close()
        logger.info("Snowflake connection closed")


if __name__ == '__main__':
    import os
    
    # Configuration
    HTML_FILE = 'queries.html'  # Change this to your HTML file path
    OUTPUT_FILE = 'incentive_output.xlsx'
    
    # Run the executor
    main(HTML_FILE, OUTPUT_FILE)
