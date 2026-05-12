"""
Excel Output Module
Handles writing query results to Excel with formatting
"""
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Writes query results to Excel with professional formatting"""
    
    def __init__(self, config: dict):
        """
        Initialize Excel writer
        
        Args:
            config: Configuration dictionary with excel settings
        """
        self.config = config
        self.excel_config = config.get('excel', {})
        self.output_file = self.excel_config.get('output_file', 'Incentive_Report.xlsx')
        self.overwrite = self.excel_config.get('overwrite', True)
        self.auto_width = self.excel_config.get('auto_width', True)
        self.freeze_header = self.excel_config.get('freeze_header', True)
        
        # Header styling
        header_style = self.excel_config.get('header_style', {})
        self.header_font = Font(
            bold=header_style.get('font', {}).get('bold', True),
            color=header_style.get('font', {}).get('color', 'FFFFFF')
        )
        self.header_fill = PatternFill(
            fill_type=header_style.get('fill', {}).get('type', 'solid'),
            start_color=header_style.get('fill', {}).get('color', '4472C4'),
            end_color=header_style.get('fill', {}).get('color', '4472C4')
        )
    
    def write_results(self, execution_results: List[Dict[str, Any]], 
                     query_config: List[Dict[str, str]]) -> str:
        """
        Write query results to Excel file
        
        Args:
            execution_results: List of execution result dictionaries
            query_config: Original query configuration with sheet names
            
        Returns:
            Path to the created Excel file
        """
        logger.info(f"Writing results to Excel: {self.output_file}")
        
        # Create mapping of query name to sheet name
        query_to_sheet = {q['name']: q['sheet_name'] for q in query_config}
        
        # Create Excel writer
        output_path = Path(self.output_file)
        
        try:
            # If overwrite is False, load existing workbook
            if not self.overwrite and output_path.exists():
                book = load_workbook(output_path)
                writer = pd.ExcelWriter(output_path, engine='openpyxl')
                writer.book = book
            else:
                writer = pd.ExcelWriter(output_path, engine='openpyxl')
            
            # Write each successful query result
            for result in execution_results:
                if result['status'] == 'success' and result['data'] is not None:
                    query_name = result['query_name']
                    sheet_name = query_to_sheet.get(query_name, query_name)
                    df = result['data']
                    
                    # Remove sheet if it exists (for overwrite)
                    if sheet_name in writer.book.sheetnames:
                        del writer.book[sheet_name]
                    
                    # Write DataFrame to sheet
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"Written {result['row_count']} rows to sheet '{sheet_name}'")
            
            # Save and close
            writer.close()
            
            # Apply formatting
            self._apply_formatting(output_path, query_to_sheet)
            
            logger.info(f"Excel file created successfully: {output_path.absolute()}")
            return str(output_path.absolute())
            
        except Exception as e:
            logger.error(f"Failed to write Excel file: {str(e)}")
            raise
    
    def _apply_formatting(self, file_path: Path, query_to_sheet: Dict[str, str]):
        """
        Apply formatting to Excel sheets
        
        Args:
            file_path: Path to Excel file
            query_to_sheet: Mapping of query names to sheet names
        """
        try:
            book = load_workbook(file_path)
            
            for sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                
                # Apply header formatting
                if sheet.max_row >= 1:
                    for cell in sheet[1]:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                if self.auto_width:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze header row
                if self.freeze_header and sheet.max_row >= 1:
                    sheet.freeze_panes = 'A2'
                
                logger.info(f"Applied formatting to sheet '{sheet_name}'")
            
            book.save(file_path)
            
        except Exception as e:
            logger.warning(f"Failed to apply some formatting: {str(e)}")
    
    def append_to_sheet(self, sheet_name: str, df: pd.DataFrame, 
                       file_path: str = None) -> str:
        """
        Append data to an existing sheet
        
        Args:
            sheet_name: Name of the sheet to append to
            df: DataFrame to append
            file_path: Path to Excel file (uses default if not provided)
            
        Returns:
            Path to the updated Excel file
        """
        if file_path is None:
            file_path = self.output_file
        
        logger.info(f"Appending {len(df)} rows to sheet '{sheet_name}'")
        
        try:
            book = load_workbook(file_path)
            
            if sheet_name not in book.sheetnames:
                # Create new sheet if it doesn't exist
                df.to_excel(file_path, sheet_name=sheet_name, index=False, 
                           mode='a', engine='openpyxl', if_sheet_exists='new')
            else:
                # Append to existing sheet
                sheet = book[sheet_name]
                start_row = sheet.max_row + 1
                
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 
                                           start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
            
            book.save(file_path)
            logger.info(f"Successfully appended to sheet '{sheet_name}'")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Failed to append to sheet: {str(e)}")
            raise
